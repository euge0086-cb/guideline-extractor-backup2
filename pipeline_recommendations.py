"""
PIPELINE DE RECOMENDACIONES — Módulo independiente
====================================================
Extrae, clasifica y exporta las recomendaciones de guías clínicas en PDF.
Compatible con ACC/AHA (2014+) y ESC (2015+).

No modifica ni depende de pipeline.py.

NOTA SOBRE TEXTO EN PDFs CON FUENTE COMPRIMIDA:
  Algunas guías (p.ej. ACC/AHA 2014) usan fuentes con "word spacing" implícito
  que pdfplumber no recupera, produciendo texto concatenado sin espacios.
  El módulo aplica corrección automática (camelCase split + wordninja si disponible).
  La Clase y el LOE se extraen correctamente; el texto puede ser imperfecto.
"""

import re
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# wordninja: opcional, mejora espaciado en texto comprimido (pip install wordninja)
try:
    import wordninja as _wordninja
    _HAS_WORDNINJA = True
except ImportError:
    _HAS_WORDNINJA = False

# Términos médicos/abreviaciones que wordninja fragmenta incorrectamente
# Comprobados contra wordninja: "troponin" → "tr opon in", "ecg" → "ecg" (OK en minúsc.)
_MEDICAL_NO_SEGMENT = frozenset({
    # Fragmentados por wordninja (verificados)
    'troponin', 'myoglobin', 'copeptin', 'bivalirudin', 'prasugrel',
    'ticagrelor', 'clopidogrel', 'fibrinolysis', 'fibrinogen',
    'electrocardiographic', 'echocardiography', 'atrioventricular',
    'revascularization', 'percutaneous', 'angiography', 'angioplasty',
    'phosphodiesterase', 'natriuretic', 'anticoagulation',
    'antithrombotic', 'antiplatelet', 'pharmacological', 'pharmacotherapy',
    'intracoronary', 'vasoconstriction', 'dysrhythmia',
    # Abreviaciones cardio que se truncan mal en minúsculas
    'ctni', 'ctnt', 'hsctn', 'nstemi', 'stemi', 'mace',
})


# ─────────────────────────────────────────────
# NORMALIZACIÓN
# ─────────────────────────────────────────────

def _normalize_class(raw: str) -> str:
    """
    Normaliza clase de recomendación → I / IIa / IIb / III.
    Acepta: "III: Harm", "III: No Benefit", prefijos COR/CLASS/CLASE,
    numerales arábigos 1/2A/2B/3.
    """
    t = str(raw).strip()
    t = re.split(r'[:\n]', t)[0].strip()   # tomar solo antes de ":"
    t = re.sub(r'\s+', '', t.upper())
    t = re.sub(r'^(CLASS|COR|CLASE|CLASSE|KLASSE)[:\s]*', '', t)
    if re.match(r'^III', t): return 'III'
    if re.match(r'^IIB', t): return 'IIb'
    if re.match(r'^IIA', t): return 'IIa'
    if t == 'II':             return 'IIa'
    if re.match(r'^I$', t):  return 'I'
    if t == '1':              return 'I'
    if t == '2A':             return 'IIa'
    if t == '2B':             return 'IIb'
    if t == '3':              return 'III'
    return raw.strip() or 'Desconocida'


def _normalize_loe(raw: str) -> str:
    """
    Normaliza nivel de evidencia → A / B / C / N/A.
    Acepta: B-R, B-NR → B | C-LD, C-EO → C | N/A, NA → N/A
    """
    t = str(raw).strip().upper()
    if t in ('N/A', 'NA', ''):
        return 'N/A'
    t = re.sub(r'^(LEVEL\s+OF\s+EVIDENCE|LEVEL|LOE|EVIDENCE)[:\s]*', '', t).strip()
    if t.startswith('A'): return 'A'
    if t.startswith('B'): return 'B'
    if t.startswith('C'): return 'C'
    return raw.strip() or 'Desconocida'


def _extract_cited_refs(text: str) -> str:
    """
    Extrae números de referencias citadas.
    Detecta superíndices Unicode, [1,2,3], (1-3), superíndices pegados a texto.
    """
    SUPERSCRIPT = str.maketrans('⁰¹²³⁴⁵⁶⁷⁸⁹', '0123456789')
    t = text.translate(SUPERSCRIPT)
    numbers: set[int] = set()

    for m in re.finditer(r'[\[\(](\d[\d,\s\-–]+\d|\d)[\]\)]', t):
        for part in re.split(r'[,\s]+', m.group(1)):
            part = part.strip()
            rng = re.match(r'^(\d+)[–\-](\d+)$', part)
            if rng:
                try:
                    numbers.update(range(int(rng.group(1)), int(rng.group(2)) + 1))
                except ValueError:
                    pass
            elif part.isdigit():
                numbers.add(int(part))

    for m in re.finditer(r'(?<=[a-zA-Z\.\)])(\d{1,3}(?:[,–\-]\d{1,3})*)', t):
        for part in re.split(r'[,–\-]', m.group(1)):
            if part.strip().isdigit():
                numbers.add(int(part.strip()))

    if not numbers:
        return ''
    return ', '.join(str(n) for n in sorted(numbers))


# ─────────────────────────────────────────────
# RECUPERACIÓN DE ESPACIADO
# ─────────────────────────────────────────────

# Términos médicos/científicos que wordninja fragmenta mal
_MEDICAL_TERMS = frozenset({
    'troponin', 'cardiac', 'cardio', 'cardiomyopathy', 'electrocardiographic',
    'electrocardiogram', 'percutaneous', 'angiography', 'angioplasty',
    'anticoagulation', 'antiplatelet', 'antithrombotic', 'fibrinolysis',
    'heparin', 'bivalirudin', 'prasugrel', 'ticagrelor', 'clopidogrel',
    'aspirin', 'nitroglycerin', 'morphine', 'metoprolol', 'bisoprolol',
    'nitroglycerin', 'phosphodiesterase', 'natriuretic', 'stratification',
    'echocardiography', 'coronary', 'revascularization', 'myocardial',
    'infarction', 'ischemia', 'ischemic', 'atrioventricular', 'ventricular',
    'arrhythmia', 'bradycardia', 'tachycardia', 'hypotension', 'hypertension',
    'biomarker', 'biomarkers', 'creatinine', 'hemoglobin', 'platelet',
    'thrombosis', 'thrombotic', 'embolism', 'hemorrhage', 'hemorrhagic',
    'contraindicated', 'contraindication', 'intermediate', 'noninvasive',
    'pharmacological', 'pharmacotherapy', 'prognostic', 'prognosis',
    'dysrhythmia', 'vasoconstriction', 'vasospasm', 'fibrinogen',
    'nondiagnostic', 'supplemental', 'sublingual', 'intravenous', 'intracoronary',
})


def _fix_spacing(text: str) -> str:
    """
    Recupera espacios en texto PDF con fuente comprimida (sin word-space glyphs).

    Estrategia:
      1. Si ya tiene >6% de espacios: devolver intacto.
      2. Dividir en tokens por puntuación y dígitos.
      3. Para cada token de texto puro:
         a. camelCase split (minúscula→Mayúscula): "ofACS" → "of ACS".
         b. Si wordninja disponible Y el token no contiene término médico
            conocido por fragmentar mal: aplicar segmentación de palabras.
         c. Rechazar resultado wordninja si contiene letras sueltas (ej. "s")
            o más del 35% de palabras de 1-2 letras.
    """
    if text.count(' ') / max(len(text), 1) > 0.06:
        return text

    def _segment(tok: str) -> str:
        if not tok or len(tok) <= 2:
            return tok
        if tok.isupper():          # acrónimo → preservar
            return tok
        # a. camelCase: "ofACS" → "of ACS"
        tok = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', tok)
        if not _HAS_WORDNINJA:
            return tok
        parts = tok.split()
        out = []
        for p in parts:
            if p.isupper() or len(p) <= 3:
                out.append(p)
                continue
            p_low = p.lower()
            # Saltar términos médicos que wordninja fragmenta mal
            if any(med in p_low for med in _MEDICAL_NO_SEGMENT):
                out.append(p)
                continue
            segs = _wordninja.split(p_low)
            # Rechazar si hay letras sueltas consonánticas (señal de error)
            bad = sum(1 for w in segs if len(w) == 1 and w not in ('a', 'i'))
            short = sum(1 for w in segs if len(w) <= 2)
            if bad > 0 or short / max(len(segs), 1) > 0.35:
                out.append(p)   # wordninja empeoraría: conservar original
            else:
                if p[0].isupper() and segs:
                    segs[0] = segs[0].capitalize()
                out.extend(segs)
        return ' '.join(out)

    # Dividir en tokens (separadores: puntuación, dígitos, espacios)
    parts = re.split(r'([,\.\(\)\[\]\-–/\*\s\d]+)', text)
    result = []
    for p in parts:
        if not p:
            continue
        if re.match(r'^[,\.\(\)\[\]\-–/\*\s\d]+$', p):
            result.append(p)
        else:
            result.append(_segment(p))
    return re.sub(r'\s+', ' ', ''.join(result)).strip()



# ─────────────────────────────────────────────
# PATRONES DE DETECCIÓN
# ─────────────────────────────────────────────

_REC_COL_RE   = re.compile(
    r'recommendation|recomendaci[oó]n|recommandation|raccomandazione', re.I
)
_CLASS_COL_RE = re.compile(
    r'^(class[a-z^\s]*|cor|clase|classe|crc|recommendation\s+class)$', re.I
)
_LOE_COL_RE   = re.compile(
    r'^(level[a-z^\s]*|loe|evidence|loe\s*\.|niveau|livello|beweis)$', re.I
)
_CLASS_VAL_RE = re.compile(
    r'^(iii(?:[:\s].+)?|iia|iib|ii(?![aibIAIB])|i(?![IiAaBbLl]))$', re.I
)
_LOE_VAL_RE   = re.compile(r'^[abc](?:-[a-z]+)?$', re.I)

# ── Ruido de página: líneas que deben EXCLUIRSE del texto de recomendaciones
_NOISE_LINE_RE = re.compile(
    r'^('
    # Cabeceras de tabla
    r'TABLE\s+\d|FIGURE\s+\d|'
    # Cabeceras de columna
    r'Recommendations?\s*$|COR\s*$|LOE\s*$|References?\s*$|'
    # Pies de tabla / notas
    r'ACS\s+indicates|CCB\s+indicates|LOE\s+indicates|'
    r'\*See\s+Section|\*Short|N/A,?\s+not\s+|'
    # Cabeceras de revista / página — ACC/AHA JACC
    r'JACC\s+VOL|Circulation\s+Vol|'
    r'e\d{3}\s+\w|'               # "e151 Amsterdam"
    r'\w+\s+et\s+al\.\s+(JACC|Circ|Eur)|'  # "Amsterdam et al. JACC"
    r'\d{4}\s+(AHA|ACC|ESC|DECEMBER|JANUARY|FEBRUARY|MARCH|APRIL|'
    r'MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER)|'
    r'DECEMBER|JANUARY|FEBRUARY|MARCH|APRIL|JUNE|JULY|AUGUST|'
    r'SEPTEMBER|OCTOBER|NOVEMBER|'
    r'\d{4}:\w*\d{3}|'            # "2014:e139"
    # Cabeceras ESC / EHJ
    r'Eur\s+Heart\s+J|EHJ\s+|'
    # Líneas de copyright/descarga
    r'Downloaded\s+from|©\s*20\d{2}|All\s+rights\s+reserved|'
    # Running header: "2014 AHA/ACC NSTE-ACS Guideline"
    r'\d{4}AHA|2014AHA|2013ESC|'
    r'NSTE-?ACSGuideline|NSTEACSGuideline'
    r')',
    re.I
)

# Patrones inline (Estrategia C — fallback)
_INLINE_PATTERNS = [
    re.compile(
        r'(?P<text>[A-Z][^(]{15,500}?)'
        r'\(Level\s+of\s+Evidence\s*[:\s]+(?P<loe>[ABC](?:-[A-Z]+)?)\)'
        r'\s*(?P<cls>I{1,3}|IIa|IIb|III)\b',
        re.I | re.DOTALL
    ),
    re.compile(
        r'(?P<text>[A-Z][^(]{15,500}?)'
        r'\((?:Class\s+)?(?P<cls>I{1,3}|IIa|IIb|III)'
        r'\s*[;,/]\s*(?:Level\s+of\s+Evidence[:\s]+|LOE\s*[:\s]*)?'
        r'(?P<loe>[ABC](?:-[A-Z]+)?)\)',
        re.I | re.DOTALL
    ),
]

# Patrón fin de línea (Estrategia B — principal para ACC/AHA y ESC)
_EOL_PAT = re.compile(
    r'^(?P<lead>.+?)\s+'
    r'(?P<cls>III(?:[:\s]+\S+(?:\s+\S+)?)?|IIa|IIb|I)\s+'
    r'(?P<loe>[ABC])\b'
    r'(?:\s+(?:N/?A|\([\d,\s\-–]+\)|\d[\d,\-–,\s]*))?'
    r'\s*$',
    re.IGNORECASE
)

# Configuraciones de tabla a probar en orden
_TABLE_SETTINGS = [
    {"vertical_strategy": "lines",  "horizontal_strategy": "lines"},
    {"vertical_strategy": "text",   "horizontal_strategy": "lines",
     "intersection_tolerance": 10,  "snap_tolerance": 3,
     "text_x_tolerance": 2},
    {"vertical_strategy": "text",   "horizontal_strategy": "text",
     "intersection_tolerance": 10,  "snap_tolerance": 3,
     "text_x_tolerance": 2},
]


# ─────────────────────────────────────────────
# ESTRATEGIA A: TABLAS
# ─────────────────────────────────────────────

def _try_extract_tables(page) -> list:
    for settings in _TABLE_SETTINGS:
        try:
            tables = page.extract_tables(table_settings=settings) or []
            if tables:
                return tables
        except Exception:
            continue
    return []


def _parse_table(table: list, page_num: int) -> list[dict]:
    if not table or len(table) < 2:
        return []

    header = [(str(c or '')).replace('\n', ' ').strip() for c in (table[0] or [])]

    rec_ci = class_ci = loe_ci = -1
    for ci, h in enumerate(header):
        h_clean = re.sub(r'\s+', ' ', h).strip()
        if _REC_COL_RE.search(h_clean)  and rec_ci   < 0: rec_ci   = ci
        if _CLASS_COL_RE.match(h_clean) and class_ci < 0: class_ci = ci
        if _LOE_COL_RE.match(h_clean)   and loe_ci   < 0: loe_ci   = ci

    if class_ci < 0:
        class_votes: dict[int, int] = {}
        loe_votes:   dict[int, int] = {}
        for row in table[1:]:
            for ci, cell in enumerate(row or []):
                v = str(cell or '').strip().split('\n')[0]
                if _CLASS_VAL_RE.match(v): class_votes[ci] = class_votes.get(ci, 0) + 1
                if _LOE_VAL_RE.match(v):   loe_votes[ci]   = loe_votes.get(ci, 0) + 1
        if class_votes: class_ci = max(class_votes, key=class_votes.get)
        if loe_votes:   loe_ci   = max(loe_votes,   key=loe_votes.get)

    if class_ci < 0:
        return []

    if rec_ci < 0:
        lengths: dict[int, int] = {}
        for row in table[1:]:
            for ci, cell in enumerate(row or []):
                if ci not in (class_ci, loe_ci):
                    lengths[ci] = lengths.get(ci, 0) + len(str(cell or ''))
        if lengths:
            rec_ci = max(lengths, key=lengths.get)

    if rec_ci < 0:
        return []

    recs = []
    for row in table[1:]:
        def cv(idx: int) -> str:
            if idx < 0 or not row or idx >= len(row):
                return ''
            return str(row[idx] or '').replace('\n', ' ').strip()

        text = _fix_spacing(cv(rec_ci))
        cls  = cv(class_ci)
        loe  = cv(loe_ci) if loe_ci >= 0 else ''

        if len(text) < 25:
            continue
        cls_norm = _normalize_class(cls)
        if cls_norm == 'Desconocida' and not cls:
            continue

        recs.append({
            'text':             text,
            'rec_class':        cls_norm,
            'loe':              _normalize_loe(loe),
            'references_cited': _extract_cited_refs(text),
            'page':             page_num,
            'source':           'tabla',
        })
    return recs


# ─────────────────────────────────────────────
# ESTRATEGIA B: FIN DE LÍNEA
# ─────────────────────────────────────────────

def _is_rec_text_line(line: str) -> bool:
    """
    ¿Es esta línea parte del texto de una recomendación?
    Descarta: cabeceras de página, títulos de tabla, secciones cortas.
    """
    s = line.strip()
    if not s or len(s) < 10:
        return False
    if _NOISE_LINE_RE.match(s):
        return False
    # Títulos de sección dentro de la tabla (cortos, sin puntuación interna)
    # p.ej. "Oxygen", "Nitrates", "CCBs", "Beta-adrenergic blockers"
    if re.match(r'^[A-Z][A-Za-z\s\-]{1,35}$', s) and len(s) <= 40:
        # Longitud corta Y sin puntuación de frase completa → título de sección
        if ',' not in s and '(' not in s and not s.endswith(')'):
            return False
    return True


def _extract_eol(page_text: str, page_num: int) -> list[dict]:
    """
    Extrae recomendaciones buscando 'texto... CLASS  LOE  [refs]' al final de línea.
    Acumula líneas intermedias para recomendaciones multilínea.
    Resetea el buffer en cualquier línea que no sea texto de recomendación.
    """
    recs = []
    lines = page_text.split('\n')
    buffer: list[str] = []

    for line in lines:
        stripped = line.strip()
        m = _EOL_PAT.match(stripped)

        if m:
            lead = m.group('lead').strip()
            cls  = m.group('cls').strip()
            loe  = m.group('loe').strip()

            full_parts = buffer + ([lead] if lead else [])
            full_text  = re.sub(r'\s+', ' ', ' '.join(full_parts)).strip()
            full_text  = _fix_spacing(full_text)

            if len(full_text) >= 25:
                recs.append({
                    'text':             full_text,
                    'rec_class':        _normalize_class(cls),
                    'loe':              _normalize_loe(loe),
                    'references_cited': _extract_cited_refs(stripped),
                    'page':             page_num,
                    'source':           'texto',
                })
            buffer = []   # siempre resetear tras una coincidencia

        elif _is_rec_text_line(stripped):
            buffer.append(stripped)
        else:
            buffer = []   # resetear en ruido, cabeceras, líneas vacías

    return recs


# ─────────────────────────────────────────────
# ESTRATEGIA C: PATRONES INLINE
# ─────────────────────────────────────────────

def _extract_inline(page_text: str, page_num: int) -> list[dict]:
    recs = []
    for pat in _INLINE_PATTERNS:
        for m in pat.finditer(page_text):
            raw_text = m.group('text').strip()
            if len(raw_text) > 400:
                last_stop = raw_text.rfind('. ', len(raw_text) - 380)
                if last_stop > 50:
                    raw_text = raw_text[last_stop + 2:].strip()
            if len(raw_text) < 25:
                continue
            recs.append({
                'text':             _fix_spacing(raw_text),
                'rec_class':        _normalize_class(m.group('cls')),
                'loe':              _normalize_loe(m.group('loe')),
                'references_cited': _extract_cited_refs(m.group('text')),
                'page':             page_num,
                'source':           'texto',
            })
    return recs


# ─────────────────────────────────────────────
# FUNCIÓN PRINCIPAL
# ─────────────────────────────────────────────

def extract_recommendations_from_pdf(pdf_path: str, debug: bool = False) -> list[dict]:
    """
    Extrae recomendaciones de una guía clínica en PDF.

    Orden de estrategias (elige la que más resultados produce):
      A — Tablas pdfplumber (3 configuraciones)
      B — Fin de línea: 'texto CLASS LOE [refs]' (ACC/AHA y ESC sin rejilla)
      C — Patrones inline: '(Class I; Level of Evidence: A)' y variantes
    """
    log = []
    def dbg(msg):
        log.append(msg)
        if debug: print(msg)

    dbg(f"[REC] wordninja {'disponible' if _HAS_WORDNINJA else 'no disponible (pip install wordninja)'}")

    all_table_recs:  list[dict] = []
    all_eol_recs:    list[dict] = []
    all_inline_recs: list[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        dbg(f"[REC] PDF: {len(pdf.pages)} páginas")

        for page_idx, page in enumerate(pdf.pages):
            page_num = page_idx + 1

            # A: tablas
            for table in _try_extract_tables(page):
                recs = _parse_table(table, page_num)
                all_table_recs.extend(recs)
                if recs:
                    dbg(f"[REC-A] pág {page_num}: {len(recs)} recs en tabla")

            # B+C: texto
            try:
                page_text = page.extract_text() or ''
            except Exception:
                page_text = ''

            if page_text:
                eol = _extract_eol(page_text, page_num)
                all_eol_recs.extend(eol)
                if eol:
                    dbg(f"[REC-B] pág {page_num}: {len(eol)} recs fin-de-línea")

                inl = _extract_inline(page_text, page_num)
                all_inline_recs.extend(inl)
                if inl:
                    dbg(f"[REC-C] pág {page_num}: {len(inl)} recs inline")

    dbg(f"[REC] A={len(all_table_recs)} B={len(all_eol_recs)} C={len(all_inline_recs)}")

    candidates = [
        (all_table_recs,  'A-tablas'),
        (all_eol_recs,    'B-eol'),
        (all_inline_recs, 'C-inline'),
    ]
    recommendations, strategy = max(candidates, key=lambda x: len(x[0]))

    if len(recommendations) < 5:
        combined = all_eol_recs + all_inline_recs
        if len(combined) > len(recommendations):
            recommendations = combined
            strategy = 'B+C combinadas'

    dbg(f"[REC] Estrategia: {strategy} → {len(recommendations)}")

    # Deduplicar
    seen: set[str] = set()
    unique: list[dict] = []
    for r in recommendations:
        key = re.sub(r'\s+', '', r['text'][:60]).lower()
        if key not in seen:
            seen.add(key)
            unique.append(r)

    for i, r in enumerate(unique, 1):
        r['rec_number'] = i

    dbg(f"[REC] Total (sin duplicados): {len(unique)}")
    if not debug:
        for line in log: print(line)

    return unique


# ─────────────────────────────────────────────
# EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────

_CLASS_COLORS = {'I': 'E2EFDA', 'IIa': 'FFF2CC', 'IIb': 'FCE4D6', 'III': 'F8CECC'}
_LOE_COLORS   = {'A': 'DAE8FC', 'B': 'E1D5E7', 'C': 'F5F5F5'}
_HEADER_BG    = '1F4E79'


def _style_header(cell, bg: str = _HEADER_BG, fg: str = 'FFFFFF'):
    cell.font      = Font(bold=True, color=fg, name='Arial', size=10)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _thin_border(ws, row: int, n_cols: int):
    thin = Side(style='thin', color='CCCCCC')
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).border = Border(bottom=thin)


def _write_rec_sheet(ws, recommendations: list[dict]):
    ws.title = 'Recomendaciones'
    cols = [
        ('Nº', 5), ('Texto de la recomendación', 75), ('Clase', 10),
        ('LOE', 12), ('Referencias citadas', 25), ('Página', 8),
        ('Fuente', 10), ('Clase (manual)', 12), ('LOE (manual)', 12), ('Notas', 30),
    ]
    ws.row_dimensions[1].height = 30
    for ci, (name, width) in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=name)
        _style_header(c)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = 'A2'

    for offset, rec in enumerate(recommendations):
        rn  = offset + 2
        cls = rec.get('rec_class', '')
        loe = rec.get('loe', '')
        values = [
            rec.get('rec_number', ''), rec.get('text', ''),
            cls, loe, rec.get('references_cited', ''),
            rec.get('page', ''), rec.get('source', ''), '', '', '',
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=rn, column=ci, value=val)
            cell.font      = Font(name='Arial', size=9)
            cell.alignment = Alignment(vertical='top', wrap_text=(ci == 2))
            if ci == 2:
                cell.fill = PatternFill('solid', start_color=_CLASS_COLORS.get(cls, 'F5F5F5'))
            elif ci == 3:
                cell.fill      = PatternFill('solid', start_color=_CLASS_COLORS.get(val, 'F5F5F5'))
                cell.alignment = Alignment(horizontal='center', vertical='top')
                cell.font      = Font(name='Arial', size=9, bold=True)
            elif ci == 4:
                cell.fill      = PatternFill('solid', start_color=_LOE_COLORS.get(val, 'F5F5F5'))
                cell.alignment = Alignment(horizontal='center', vertical='top')
                cell.font      = Font(name='Arial', size=9, bold=True)
        ws.row_dimensions[rn].height = 55
        _thin_border(ws, rn, len(cols))


def _write_summary_sheet(ws, recommendations: list[dict]):
    ws.title = 'Resumen recomendaciones'
    CLASS_ORDER = ['I', 'IIa', 'IIb', 'III', 'Desconocida']
    LOE_ORDER   = ['A', 'B', 'C', 'N/A', 'Desconocida']

    matrix:       dict[str, dict[str, int]] = {}
    class_totals: dict[str, int]            = {}
    loe_totals:   dict[str, int]            = {}
    for rec in recommendations:
        cls = rec.get('rec_class', 'Desconocida')
        loe = rec.get('loe', 'Desconocida')
        matrix.setdefault(cls, {})
        matrix[cls][loe]  = matrix[cls].get(loe, 0) + 1
        class_totals[cls] = class_totals.get(cls, 0) + 1
        loe_totals[loe]   = loe_totals.get(loe, 0)  + 1

    present_cls = [c for c in CLASS_ORDER if c in class_totals]
    present_loe = [l for l in LOE_ORDER   if l in loe_totals]
    ws.column_dimensions['A'].width = 20

    _style_header(ws.cell(row=1, column=1, value='Clase / LOE'))
    for ci, loe in enumerate(present_loe, 2):
        _style_header(ws.cell(row=1, column=ci, value=loe))
        ws.column_dimensions[get_column_letter(ci)].width = 10
    total_col = len(present_loe) + 2
    _style_header(ws.cell(row=1, column=total_col, value='TOTAL'))
    ws.column_dimensions[get_column_letter(total_col)].width = 10

    for ri, cls in enumerate(present_cls, 2):
        c = ws.cell(row=ri, column=1, value=cls)
        c.font = Font(bold=True, name='Arial', size=10)
        c.fill = PatternFill('solid', start_color=_CLASS_COLORS.get(cls, 'F5F5F5'))
        for ci, loe in enumerate(present_loe, 2):
            val = matrix.get(cls, {}).get(loe, 0)
            cell = ws.cell(row=ri, column=ci, value=val if val else '')
            cell.font      = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center')
        t = ws.cell(row=ri, column=total_col, value=class_totals[cls])
        t.font = Font(bold=True, name='Arial', size=10)
        t.alignment = Alignment(horizontal='center')

    tr = len(present_cls) + 2
    ws.cell(row=tr, column=1, value='TOTAL').font = Font(bold=True, name='Arial', size=10)
    for ci, loe in enumerate(present_loe, 2):
        c = ws.cell(row=tr, column=ci, value=loe_totals.get(loe, 0))
        c.font = Font(bold=True, name='Arial', size=10)
        c.alignment = Alignment(horizontal='center')
    g = ws.cell(row=tr, column=total_col, value=len(recommendations))
    g.font = Font(bold=True, name='Arial', size=10)
    g.alignment = Alignment(horizontal='center')

    leyenda = [
        ('CLASE I',   'Beneficio >> riesgo. Indicado / recomendado.',            'E2EFDA'),
        ('CLASE IIa', 'Beneficio > riesgo. Es razonable realizarlo.',             'FFF2CC'),
        ('CLASE IIb', 'Beneficio ≥ riesgo. Puede considerarse.',                  'FCE4D6'),
        ('CLASE III', 'Sin beneficio o dañino. No está recomendado.',             'F8CECC'),
        ('LOE A',     'Múltiples ECAs o meta-análisis de alta calidad.',          'DAE8FC'),
        ('LOE B',     'Un ECA o estudios no aleatorizados de gran tamaño.',       'E1D5E7'),
        ('LOE C',     'Consenso de expertos, estudios pequeños o registros.',     'F5F5F5'),
    ]
    for offset, (label, desc, color) in enumerate(leyenda, tr + 2):
        c = ws.cell(row=offset, column=1, value=label)
        c.font = Font(name='Arial', size=9)
        c.fill = PatternFill('solid', start_color=color)
        ws.cell(row=offset, column=2, value=desc).font = Font(name='Arial', size=9)
    ws.column_dimensions['B'].width = 55


def export_recommendations_to_excel(recommendations: list[dict], output_path: str):
    wb = Workbook()
    _write_rec_sheet(wb.active, recommendations)
    _write_summary_sheet(wb.create_sheet(), recommendations)
    wb.save(output_path)
    print(f'[OK] Excel guardado: {output_path}')
