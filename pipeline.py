"""
GUIDELINE REFERENCE PIPELINE
================================
Pipeline modular para extraer referencias de guías clínicas en PDF,
enriquecer con metadatos (PubMed + CrossRef) y exportar a Excel.

MÓDULOS:
  1. extract_references(pdf_path)  → lista de strings de referencias crudas
  2. enrich_reference(ref_text)    → dict con PMID, DOI, autores, año, etc.
  3. classify_reference(metadata)  → tipo: RCT_primario / RCT_secundario / meta-analisis / observacional / otro
  4. export_to_excel(records, out) → archivo .xlsx con hojas diferenciadas

USO:
  python guideline_pipeline.py <ruta_al_pdf> [output.xlsx]
"""

import re
import sys
import time
import json
import requests
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# MÓDULO 1: EXTRACCIÓN DE REFERENCIAS DEL PDF
# ─────────────────────────────────────────────

def extract_references_from_pdf(pdf_path: str, debug: bool = False) -> list[str]:
    """
    Extrae TODAS las referencias de un PDF (artículo o guía clínica),
    de forma UNIVERSAL e independiente del layout editorial.

    ESTRATEGIA EN DOS FASES:

    FASE 1 — Punto de inicio universal: se localiza la ÚLTIMA aparición
    del header "References" en todo el documento. La sección de
    referencias real de cualquier artículo o guía es siempre la última
    (evita falsos positivos de menciones previas en el índice o cuerpo
    del texto). A partir de esa página se considera que todo lo que
    sigue son referencias.

    FASE 2 — Reconstrucción robusta dentro de ese bloque: en vez de
    asumir la geometría de columnas (frágil y específica de cada
    editorial), se localizan los marcadores numéricos de inicio de
    referencia ("1.", "01.", "827.") que aparecen como PRIMERA palabra
    de su línea. Se valida la secuencia con el algoritmo de subsecuencia
    creciente más larga (LIS), y las posiciones X de esos marcadores se
    clusterizan dinámicamente para inferir cuántas columnas tiene el
    documento — sin asumir ningún ancho o número fijo.
    """
    log = []

    def dbg(msg):
        log.append(msg)
        if debug:
            print(msg)

    HEADER_PATTERNS = [
        r'^references?\s*:?\s*\d{0,4}\s*$',
        r'^\d{1,3}\s+references?\s*$',          # "17 References"
        r'^bibliograf[ií]a\s*:?\s*\d{0,4}\s*$',
        r'^referenci[ae]s\s*:?\s*\d{0,4}\s*$',
        r'^literature\s+cited\s*$',
        r'^works\s+cited\s*$',
    ]
    SECTION_END_PATTERNS = [
        r'^table\s+[a-z0-9]+\b', r'^figure\s+\d+',
        r'^appendix\b', r'^central\s+illustration\b',
    ]
    NOISE_PATTERNS = [
        r'^(jacc|circulation|eur\s*heart\s*j?|j\s*am\s*coll\s*cardiol)\b.{0,50}\d{4}',
        r'^\d{4}\s*(aha|acc|esc)\b.{0,80}guideline',
        r'^downloaded\s+from', r'page\s+\d+\s+of\s+\d+',
        r'^esc\s+guidelines\s*\d{0,5}$',          # pie de página típico ESC
        r'^\d{4}\s*esc\s+guidelines\s*$',
    ]

    def is_header(line: str) -> bool:
        c = line.strip()
        return any(re.match(p, c, re.IGNORECASE) for p in HEADER_PATTERNS)

    def is_noise(text: str) -> bool:
        return any(re.search(p, text, re.IGNORECASE) for p in NOISE_PATTERNS)

    # Acepta "1." y también "01." (ceros a la izquierda, frecuente en ESC)
    MARKER_RE = re.compile(r'^0*(\d{1,4})[\.\)](.*)$')

    def looks_like_reference_start(rest_text: str, next_word_text: str = "") -> bool:
        combined = (rest_text + " " + next_word_text).strip()
        if not combined:
            return False
        if not combined[0].isupper():
            return False
        first_tok = combined.split()[0] if combined.split() else ""
        if re.match(r'^\d{4}\.?$', first_tok):
            return False
        if re.match(r'^(https?|doi|www)', combined, re.IGNORECASE):
            return False
        return True

    # ── FASE 1: localizar la ÚLTIMA aparición de "References" ────────────
    all_words_by_page = []
    header_pages = []

    with pdfplumber.open(pdf_path) as pdf:
        n_pages = len(pdf.pages)
        dbg(f"[INFO] PDF con {n_pages} páginas")

        for page_idx, page in enumerate(pdf.pages):
            try:
                words = page.extract_words()
            except Exception:
                words = []
            all_words_by_page.append(words)

            line_map = {}
            for w in words:
                y_key = round(w["top"])
                line_map.setdefault(y_key, []).append(w["text"])
            for toks in line_map.values():
                line = " ".join(toks).strip()
                if is_header(line):
                    header_pages.append(page_idx)
                    break

    dbg(f"[INFO] Apariciones de header 'References' en el documento: "
        f"{len(header_pages)} → páginas {[p+1 for p in header_pages]}")

    if header_pages:
        last_header_page = header_pages[-1]
        dbg(f"[INFO] Usando la ÚLTIMA aparición (página {last_header_page + 1}) "
            f"como inicio de la sección de referencias")
        candidate_pages = set(range(last_header_page, n_pages))
    else:
        dbg("[WARN] No se encontró ningún header 'References'. "
            "Se buscarán marcadores numerados en todo el documento.")
        candidate_pages = set(range(n_pages))

    # ── FASE 2: localizar marcadores de inicio de referencia ─────────────
    # IMPORTANTE: NO agrupamos primero por línea-Y combinando toda la
    # página, porque dos columnas distintas pueden compartir la misma
    # coordenada Y (su texto está en la misma fila visual) y eso fusiona
    # erróneamente ambas columnas en una sola "línea", perdiendo el
    # marcador de la columna derecha. En su lugar, evaluamos cada
    # palabra de forma independiente como posible marcador, y para
    # validarla buscamos la palabra que le sigue espacialmente DENTRO
    # DE SU PROPIA COLUMNA (incremento de Y, X similar) — no la
    # siguiente palabra en X dentro de la misma fila combinada.
    all_markers = []
    for page_idx in sorted(candidate_pages):
        page_words = all_words_by_page[page_idx]

        for w in page_words:
            m = MARKER_RE.match(w["text"])
            if not m:
                continue
            num = int(m.group(1))
            rest = m.group(2)

            # Si el propio token ya trae texto pegado (p.ej. "5.Bonow"),
            # usamos eso. Si no (p.ej. "5." y "Bonow" son tokens
            # separados), buscamos la palabra más cercana que SIGA al
            # marcador en su misma columna: misma fila (top similar,
            # x0 inmediatamente mayor) o, si no hay nada en esa fila,
            # la siguiente fila por debajo con x0 similar al marcador.
            next_word_text = ""
            if not rest.strip():
                same_row_candidates = [
                    ww for ww in page_words
                    if abs(ww["top"] - w["top"]) < 1.5 and ww["x0"] > w["x1"] and ww["x0"] - w["x1"] < 15
                ]
                if same_row_candidates:
                    same_row_candidates.sort(key=lambda ww: ww["x0"])
                    next_word_text = same_row_candidates[0]["text"]
                else:
                    below_candidates = [
                        ww for ww in page_words
                        if ww["top"] > w["top"] + 1.5
                        and abs(ww["x0"] - w["x0"]) < 30
                    ]
                    if below_candidates:
                        below_candidates.sort(key=lambda ww: ww["top"])
                        next_word_text = below_candidates[0]["text"]

            if not looks_like_reference_start(rest, next_word_text):
                continue

            all_markers.append({
                "num": num, "page": page_idx,
                "x0": w["x0"], "x1": w["x1"], "top": w["top"],
                "rest": rest,
            })

    dbg(f"[INFO] Marcadores numéricos candidatos: {len(all_markers)}")

    if not all_markers:
        dbg("[ERROR] No se encontraron marcadores de referencia numerados.")
        if not debug:
            for l in log:
                print(l)
        return []

    # ── Validar la secuencia mediante LIS (Longest Increasing Subsequence)
    all_markers.sort(key=lambda m: m["num"])
    nums = [m["num"] for m in all_markers]
    n = len(nums)

    dp = [1] * n
    parent = [-1] * n
    for i in range(n):
        for j in range(i):
            if nums[j] < nums[i] and dp[j] + 1 > dp[i]:
                dp[i] = dp[j] + 1
                parent[i] = j
    best_end = max(range(n), key=lambda i: dp[i])
    lis_indices = []
    cur = best_end
    while cur != -1:
        lis_indices.append(cur)
        cur = parent[cur]
    lis_indices.reverse()

    valid_markers = [all_markers[i] for i in lis_indices]
    valid_markers.sort(key=lambda m: (m["page"], m["top"], m["x0"]))

    dbg(f"[INFO] Marcadores validados (secuencia creciente más larga): "
        f"{len(valid_markers)} de {len(all_markers)} candidatos "
        f"(#{valid_markers[0]['num'] if valid_markers else '-'} a "
        f"#{valid_markers[-1]['num'] if valid_markers else '-'})")

    if len(valid_markers) < 3:
        dbg("[ERROR] Menos de 3 marcadores válidos en secuencia. Abortando.")
        if not debug:
            for l in log:
                print(l)
        return []

    # ── Clusterizar posiciones X de los marcadores POR PÁGINA (no
    # globalmente) para inferir columnas. Esto es necesario porque
    # algunos PDFs alternan el margen entre páginas pares/impares
    # (formato libro: la columna "izquierda" puede estar en x≈42 en
    # una página y en x≈72 en la siguiente), lo que rompería un
    # clustering global de coordenadas X absolutas, fragmentando
    # columnas reales en clusters falsos adicionales.
    #
    # En su lugar, para cada página se determina cuántas columnas
    # tiene y se asigna a cada marcador un ÍNDICE de columna relativo
    # a esa página (0 = más a la izquierda, 1 = siguiente, ...). Ese
    # índice es estable entre páginas aunque el desplazamiento de
    # margen cambie.
    markers_by_page = {}
    for m in valid_markers:
        markers_by_page.setdefault(m["page"], []).append(m)

    page_column_centers = {}  # page_idx -> [x0_centro_col0, x0_centro_col1, ...]
    for page_idx, page_markers in markers_by_page.items():
        xs = sorted(set(round(m["x0"]) for m in page_markers))
        clusters = []
        for x in xs:
            placed = False
            for cluster in clusters:
                if abs(x - cluster[-1]) < 20:
                    cluster.append(x)
                    placed = True
                    break
            if not placed:
                clusters.append([x])
        centers = sorted(sum(c) / len(c) for c in clusters)
        page_column_centers[page_idx] = centers

    n_columns_detected = max((len(c) for c in page_column_centers.values()), default=1)
    dbg(f"[INFO] Columnas detectadas (máximo por página, recalculado página a página): "
        f"{n_columns_detected}")

    def column_index(page_idx, x0):
        """Devuelve el índice de columna (0, 1, 2...) de una posición X
        dentro de la página dada, usando los centros de columna ya
        calculados para esa página específica."""
        centers = page_column_centers.get(page_idx)
        if not centers:
            return 0
        return min(range(len(centers)), key=lambda i: abs(centers[i] - x0))

    def nearest_column(page_idx, x0):
        return column_index(page_idx, x0)

    # ── Reconstruir el texto de cada referencia ───────────────────────────
    # Para delimitar el final de cada referencia usamos el SIGUIENTE
    # marcador que esté en su MISMA columna (no necesariamente el
    # siguiente en la secuencia global), ya que dos columnas avanzan en
    # paralelo y el marcador inmediatamente siguiente en número puede
    # estar en la columna de al lado, sin relación de continuidad visual.
    full_refs = []
    for i, mk in enumerate(valid_markers):
        mk_column = nearest_column(mk["page"], mk["x0"])

        # Buscar el siguiente marcador (en orden de número) que comparta
        # la misma columna que el actual (índice de columna relativo a
        # SU PROPIA página, no coordenada X absoluta)
        next_mk = None
        for j in range(i + 1, len(valid_markers)):
            candidate = valid_markers[j]
            if nearest_column(candidate["page"], candidate["x0"]) == mk_column:
                next_mk = candidate
                break

        collected_words = [mk["rest"]] if mk["rest"].strip() else []

        start_page = mk["page"]
        end_page = next_mk["page"] if next_mk else start_page

        for page_idx in range(start_page, end_page + 1):
            page_words = all_words_by_page[page_idx]
            for w in page_words:
                if page_idx == mk["page"] and abs(w["top"] - mk["top"]) < 0.5 and w["x0"] == mk["x0"]:
                    continue

                same_page_as_marker = (page_idx == mk["page"])
                after_start = True
                if same_page_as_marker:
                    after_start = w["top"] >= mk["top"] - 0.5

                before_end = True
                if next_mk and page_idx == next_mk["page"]:
                    if page_idx == mk["page"]:
                        before_end = (w["top"] < next_mk["top"] - 0.5) or \
                                     (abs(w["top"] - next_mk["top"]) < 0.5 and w["x0"] < next_mk["x0"])
                    else:
                        before_end = w["top"] < next_mk["top"] - 0.5

                # La columna de la palabra se calcula respecto a SU
                # PROPIA página (no la del marcador), ya que el margen
                # puede desplazarse de una página a otra
                in_column = nearest_column(page_idx, w["x0"]) == mk_column

                if after_start and before_end and in_column:
                    collected_words.append(w["text"])

            if next_mk and page_idx == next_mk["page"]:
                break

        ref_text = f"{mk['num']}. " + " ".join(collected_words)
        ref_text = re.sub(r'\s+', ' ', ref_text).strip()

        if is_noise(ref_text[:80]):
            continue

        full_refs.append(ref_text)

    # ── Filtrar referencias degeneradas ───────────────────────────────────
    def looks_like_table_row(text: str) -> bool:
        digit_ratio = sum(c.isdigit() for c in text) / max(len(text), 1)
        return digit_ratio > 0.45 or text.count('|') > 2

    clean_refs = [r for r in full_refs if len(r) >= 20 and not looks_like_table_row(r)]

    # ── Detectar huecos en la numeración final para informar al usuario.
    # En casos raros, ciertas zonas del PDF tienen texto con codificación
    # de fuente anómala (caracteres mal espaciados o superpuestos) que
    # ningún ajuste de tolerancia logra reconstruir correctamente; cuando
    # eso ocurre, esas referencias puntuales no se detectan como
    # marcador válido. Se informa explícitamente en vez de fallar en
    # silencio, para que el usuario sepa exactamente qué números faltan
    # y pueda completarlos manualmente si lo necesita.
    extracted_nums = set()
    for r in clean_refs:
        m = re.match(r'^(\d+)\.', r)
        if m:
            extracted_nums.add(int(m.group(1)))
    if extracted_nums:
        full_range = set(range(min(extracted_nums), max(extracted_nums) + 1))
        gaps = sorted(full_range - extracted_nums)
        if gaps:
            dbg(f"[AVISO] {len(gaps)} número(s) de referencia no se pudieron "
                f"extraer dentro del rango {min(extracted_nums)}-{max(extracted_nums)}: "
                f"{gaps}. Esto suele deberse a zonas puntuales del PDF con "
                f"codificación de fuente irregular (caracteres mal espaciados). "
                f"Revisa esas referencias manualmente en el PDF original.")

    dbg(f"[INFO] Referencias extraídas (total final, tras limpieza): {len(clean_refs)}")
    if not debug:
        for l in log:
            print(l)

    return clean_refs


# ─────────────────────────────────────────────
# MÓDULO 2: ENRIQUECIMIENTO VÍA APIs
# ─────────────────────────────────────────────

def search_pubmed(query: str, api_key: str = "") -> dict:
    """Busca en PubMed: PMID + metadatos + Publication Types oficiales."""
    base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    params = {"db": "pubmed", "term": query, "retmax": 1, "retmode": "json"}
    if api_key: params["api_key"] = api_key
    try:
        r = requests.get(f"{base}esearch.fcgi", params=params, timeout=10)
        ids = r.json().get("esearchresult", {}).get("idlist", [])
        if not ids: return {}
        pmid = ids[0]
        fetch_params = {"db": "pubmed", "id": pmid, "retmode": "xml", "rettype": "abstract"}
        if api_key: fetch_params["api_key"] = api_key
        rf = requests.get(f"{base}efetch.fcgi", params=fetch_params, timeout=10)
        xml = rf.text
        def extract_xml(tag, text):
            m = re.search(rf'<{tag}[^>]*>(.*?)</{tag}>', text, re.DOTALL)
            return m.group(1).strip() if m else ""
        title   = extract_xml("ArticleTitle", xml)
        year    = extract_xml("Year", xml) or extract_xml("MedlineDate", xml)[:4]
        journal = extract_xml("Title", xml)
        authors_raw = re.findall(r'<LastName>(.*?)</LastName>.*?<ForeName>(.*?)</ForeName>', xml, re.DOTALL)
        authors = ", ".join([f"{ln} {fn[0]}." for ln, fn in authors_raw[:3]])
        if len(authors_raw) > 3: authors += " et al."
        doi_m = re.search(r'<ArticleId IdType="doi">(.*?)</ArticleId>', xml)
        doi = doi_m.group(1).strip() if doi_m else ""
        pub_types  = re.findall(r'<PublicationType[^>]*>(.*?)</PublicationType>', xml)
        mesh_terms = re.findall(r'<DescriptorName[^>]*>(.*?)</DescriptorName>', xml)
        # El abstract puede venir en uno o varios bloques <AbstractText>
        # (algunos artículos lo dividen en secciones: Background,
        # Methods, Results, Conclusions, cada una con su propia etiqueta)
        abstract_blocks = re.findall(r'<AbstractText[^>]*>(.*?)</AbstractText>', xml, re.DOTALL)
        abstract = " ".join(b.strip() for b in abstract_blocks) if abstract_blocks else ""
        # Limpiar posibles tags HTML residuales dentro del abstract (p.ej. <i>, <sub>)
        abstract = re.sub(r'<[^>]+>', '', abstract).strip()
        return {"pmid": pmid, "doi": doi, "title": title,
                "year": year[:4] if year else "", "journal": journal,
                "authors": authors, "pub_types": pub_types,
                "mesh_terms": mesh_terms, "abstract": abstract,
                "source": "PubMed"}
    except:
        return {}


def fetch_pubtypes_by_pmid(pmid: str, api_key: str = "") -> tuple:
    """Obtiene Publication Types y MeSH de PubMed dado un PMID ya conocido."""
    base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
    params = {"db": "pubmed", "id": pmid, "retmode": "xml", "rettype": "abstract"}
    if api_key: params["api_key"] = api_key
    try:
        r = requests.get(base, params=params, timeout=10)
        pub_types  = re.findall(r'<PublicationType[^>]*>(.*?)</PublicationType>', r.text)
        mesh_terms = re.findall(r'<DescriptorName[^>]*>(.*?)</DescriptorName>', r.text)
        abstract_blocks = re.findall(r'<AbstractText[^>]*>(.*?)</AbstractText>', r.text, re.DOTALL)
        abstract = " ".join(b.strip() for b in abstract_blocks) if abstract_blocks else ""
        abstract = re.sub(r'<[^>]+>', '', abstract).strip()
        return pub_types, mesh_terms, abstract
    except:
        return [], [], ""


def search_crossref(ref_text: str) -> dict:
    """Busca en CrossRef por texto de referencia libre."""
    url = "https://api.crossref.org/works"
    # Limpiar número de referencia
    clean = re.sub(r'^\d+[\.\s]+', '', ref_text).strip()
    # Insertar espacios en límites camelCase (texto de PDFs con fuente
    # comprimida, sin espacios reales entre palabras)
    clean = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', clean)
    clean = clean[:200]
    params = {
        "query.bibliographic": clean,
        "rows": 1,
        "select": "DOI,title,author,published,container-title,type"
    }
    headers = {"User-Agent": "Guideline-Pipeline/1.0 (research tool)"}
    try:
        r = requests.get(url, params=params, headers=headers, timeout=10)
        items = r.json().get("message", {}).get("items", [])
        if not items:
            return {}
        item = items[0]
        
        doi = item.get("DOI", "")
        title = item.get("title", [""])[0] if item.get("title") else ""
        pub_type = item.get("type", "")
        
        authors_list = item.get("author", [])
        authors = ", ".join([
            f"{a.get('family', '')} {a.get('given', [''])[0]}." 
            if a.get('given') else a.get('family', '')
            for a in authors_list[:3]
        ])
        if len(authors_list) > 3:
            authors += " et al."
        
        year = ""
        pub = item.get("published", {}).get("date-parts", [[""]])
        if pub and pub[0]:
            year = str(pub[0][0])
        
        journal = ""
        ct = item.get("container-title", [])
        if ct:
            journal = ct[0]
        
        return {
            "doi": doi,
            "title": title,
            "year": year,
            "journal": journal,
            "authors": authors,
            "pub_type_raw": pub_type,
            "source": "CrossRef"
        }
    except Exception as e:
        return {}


def build_pubmed_query(ref_text: str) -> str:
    """Construye query PubMed desde texto de referencia."""
    # Extraer título tentativo: texto entre el primer punto y la revista
    clean = re.sub(r'^\d+[\.\s]+', '', ref_text).strip()

    # Si el texto viene de un PDF con fuente comprimida (sin espacios
    # entre palabras, p.ej. "TheJointCommission.AcuteMyocardialInfarc-"),
    # las palabras quedan pegadas y la búsqueda en PubMed falla porque
    # nunca encuentra términos reales. Insertamos espacios en los
    # límites camelCase (minúscula→Mayúscula) antes de tokenizar.
    clean = re.sub(r'(?<=[a-z])(?=[A-Z])', ' ', clean)
    clean = re.sub(r'-\s', '', clean)  # reunir palabras partidas por guion de fin de línea

    # Intentar extraer primeras palabras significativas del título
    words = re.findall(r'\b[A-Za-z]{4,}\b', clean)
    query = " ".join(words[:8])
    # Añadir año si está presente
    year_m = re.search(r'\b(19|20)\d{2}\b', clean)
    if year_m:
        query += f"[Title/Abstract] AND {year_m.group()}[PDAT]"
    return query


def enrich_reference(ref_text: str, idx: int) -> dict:
    """
    Enriquece una referencia con metadatos desde PubMed y CrossRef.
    Combina los resultados priorizando PubMed para PMID y CrossRef para DOI.

    IMPORTANTE: el número de referencia mostrado ("N°") se extrae del
    propio texto (ref_text), que el extractor ya graba con su número
    real (p.ej. "825. TheJointCommission..."), NO del índice de
    enumeración (idx) de la lista en memoria. Esto evita que el "N°"
    del Excel quede desincronizado del número real si la lista de
    referencias no llega perfectamente ordenada (p.ej. tramos finales
    de documentos con layout irregular, páginas de créditos, etc.).
    El parámetro idx se conserva solo como fallback si no se puede
    extraer el número real del texto.
    """
    real_num_match = re.match(r'^\s*0*(\d{1,4})[\.\)]', ref_text)
    ref_number = int(real_num_match.group(1)) if real_num_match else idx

    record = {
        "ref_number": ref_number,
        "ref_raw": ref_text,
        "pmid": "",
        "doi": "",
        "title": "",
        "authors": "",
        "year": "",
        "journal": "",
        "abstract": "",
        "study_type": "",
        "study_type_auto": "",
        "classification_criterion": "",
        "pubmed_url": "",
        "doi_url": "",
        "source_api": "",
        "pub_types": [],    # Publication Types oficiales de PubMed
        "mesh_terms": [],   # MeSH terms de PubMed
        "pub_type_raw": "", # Tipo CrossRef
        "notes": ""
    }

    # Detectar referencias claramente NO académicas (páginas web
    # institucionales, "Available at: http://...", sin DOI/PMID
    # esperable) para no malgastar búsquedas que sabemos que fallarán
    # y producen falsos positivos en PubMed/CrossRef.
    is_web_source = bool(re.search(
        r'available\s*at\s*:?\s*https?://|^\s*\d+\.\s*[A-Za-z\s]+\.\s*(20\d{2}|19\d{2})\.\s*Available',
        ref_text, re.IGNORECASE
    ))
    if is_web_source:
        record["notes"] = "Fuente web/institucional (no indexada en PubMed/CrossRef)"
        record["study_type_auto"] = "fuente_web/institucional"
        record["classification_criterion"] = "Detección por patrón de cita web/institucional ('Available at: http://...') antes de cualquier búsqueda PubMed/CrossRef"
        url_m = re.search(r'(https?://\S+)', ref_text)
        if url_m:
            record["doi_url"] = url_m.group(1).rstrip('.')
        return record

    time.sleep(0.35)  # Respetar rate limit NCBI (3 req/s sin API key)

    # 1. CrossRef primero (más tolerante a texto libre)
    cr = search_crossref(ref_text)
    if cr:
        for k, v in cr.items():
            if v and k in record:
                record[k] = v
        record["source_api"] = "CrossRef"

    # 2. PubMed para PMID + Publication Types (clave para clasificación)
    query = build_pubmed_query(ref_text)
    if query:
        time.sleep(0.35)
        pm = search_pubmed(query)
        if pm:
            if pm.get("pmid"):
                record["pmid"] = pm["pmid"]
                record["pubmed_url"] = f"https://pubmed.ncbi.nlm.nih.gov/{pm['pmid']}/"
            for k in ["doi", "title", "authors", "year", "journal", "abstract"]:
                if not record[k] and pm.get(k):
                    record[k] = pm[k]
            # Guardar Publication Types y MeSH para el clasificador
            record["pub_types"]  = pm.get("pub_types", [])
            record["mesh_terms"] = pm.get("mesh_terms", [])
            record["source_api"] = "PubMed+CrossRef" if cr else "PubMed"

    # 3. Si ya tenemos PMID pero no pub_types (vino solo de CrossRef), buscarlos
    if record["pmid"] and not record["pub_types"]:
        time.sleep(0.35)
        pt, mt, ab = fetch_pubtypes_by_pmid(record["pmid"])
        record["pub_types"]  = pt
        record["mesh_terms"] = mt
        if not record["abstract"] and ab:
            record["abstract"] = ab

    # 4. Si tenemos PMID pero por algún motivo el abstract sigue vacío
    # (p.ej. PubMed search no devolvió abstract en la primera pasada),
    # intentar recuperarlo explícitamente una última vez
    if record["pmid"] and not record["abstract"]:
        time.sleep(0.35)
        _, _, ab = fetch_pubtypes_by_pmid(record["pmid"])
        if ab:
            record["abstract"] = ab

    if record["doi"]:
        record["doi_url"] = f"https://doi.org/{record['doi']}"

    return record


# ─────────────────────────────────────────────
# MÓDULO 3: CLASIFICACIÓN AUTOMÁTICA
# ─────────────────────────────────────────────

# Palabras clave para clasificación por tipo de estudio
CLASSIFICATION_RULES = {
    "RCT_primario": [
        r'\brandomis[ei]d\b', r'\bplacebo.controlled\b',
        r'\bdouble.blind\b', r'\bsingle.blind\b',
        r'\brandom(ized|ised)\s+(clinical|controlled)\s+trial\b',
        r'\bRCT\b', r'\bensayo\s+cl[ií]nico\b', r'\brandomizado\b',
        r'\bprimary\s+(result|endpoint|outcome)\b',
        r'\btrial\b.*\bplacebo\b', r'\bversus\b.*\bplacebo\b',
    ],
    "RCT_secundario": [
        r'\bsubgroup\s+anal', r'\bpost.hoc\b', r'\bsecondary\s+anal',
        r'\bsub-?study\b', r'\bpre.specified\b', r'\bpost\s+hoc\b',
        r'\bsubstudy\b', r'\bsub\s+analysis\b',
        r'\binsights\s+from\b', r'\bobservations\s+from\b',
        r'\ba\s+substudy\s+of\b', r'\bsecondary\s+(endpoint|outcome)\b',
        r'\bpredictors?\s+of\s+outcome\b', r'\bexploratory\s+analysis\b',
    ],
    "meta-analisis": [
        r'\bmeta.anal', r'\bsystematic\s+review\b', r'\bpooled\s+anal',
        r'\bsystematic\b.*\breview\b', r'\bmetaan[aá]lisis\b',
        r'\bindividual\s+patient\s+data\b', r'\bnetwork\s+meta\b',
    ],
    "registro_observacional": [
        r'\bregist(ry|er|ro)\b', r'\bcohort\b', r'\bobservational\b',
        r'\bretrospective\b', r'\bprospective\s+(cohort|observational)\b',
        r'\bepidemiolog\b', r'\bsurvey\b', r'\bpopulation.based\b',
        r'\bdatabase\b', r'\bcross.sectional\b',
        # Registros conocidos en texto crudo
        r'\bGRACE\b', r'\bSWEDEHEART\b', r'\bNRMI\b', r'\bCRUSADE\b',
        r'\bACTION\b.*\bregist', r'\bEHS\b', r'\bEURO\s*HEART\b',
        r'\bNHANES\b', r'\bFRAMINGHAM\b',
    ],
    "guia_clinica": [
        r'\bguidelines?\b', r'\brecommendations?\b',
        r'\bconsensus\s+(statement|document|report)\b',
        r'\bgu[ií]a\s+(cl[ií]nica|de\s+pr[aá]ctica)\b',
        r'\btask\s+force\b', r'\bwriting\s+(committee|group)\b',
        r'\bpractice\s+guideline\b', r'\bexpert\s+consensus\b',
        r'\bposition\s+(statement|paper)\b',
        r'\bESC\s+guideline', r'\bACC.AHA\s+guideline',
        r'\bACCF.AHA\s+guideline', r'\bAHA.ACC\s+guideline',
        r'\bfocused\s+update\b', r'\bpolicy\s+statement\b',
        r'\bscientific\s+statement\b',
    ],
}

# Mapeo de Publication Types de PubMed → categorías de la app
PUBMED_TYPE_MAP = {
    # ECA primario
    "Randomized Controlled Trial":              "RCT_primario",
    "Controlled Clinical Trial":                "RCT_primario",
    "Clinical Trial, Phase III":                "RCT_primario",
    "Clinical Trial, Phase IV":                 "RCT_primario",
    "Multicenter Study":                        None,  # complementario, no definitivo
    # ECA secundario
    "Clinical Trial":                           None,  # demasiado genérico solo
    # Meta-análisis
    "Meta-Analysis":                            "meta-analisis",
    "Systematic Review":                        "meta-analisis",
    # Observacional / registro
    "Observational Study":                      "registro_observacional",
    "Multicenter Study":                        None,
    # Guía clínica
    "Practice Guideline":                       "guia_clinica",
    "Guideline":                                "guia_clinica",
    "Consensus Development Conference":         "guia_clinica",
    "Consensus Development Conference, NIH":    "guia_clinica",
    "Government Publications":                  None,
}

def classify_reference(record: dict) -> tuple:
    """
    Clasificación en 3 capas de precisión decreciente:
    1. Publication Types oficiales de PubMed (más fiable)
    2. Palabras clave en título + texto crudo + journal
    3. Fallback por tipo CrossRef + vocabulario de intervención

    Devuelve una tupla (study_type, criterio) donde 'criterio' es una
    descripción legible de la regla exacta que disparó la clasificación,
    para que sea auditable caso por caso (columna 'Criterio' en el Excel).
    """

    # ── CAPA 1: Publication Types de PubMed ─────────────────────────────
    pub_types = record.get("pub_types", [])  # lista de strings de PubMed
    if pub_types:
        type_str = " | ".join(pub_types)

        # Subanálisis: Clinical Trial + sin "Randomized" = probable secundario
        if any("Randomized" in t for t in pub_types):
            if any(kw in type_str for kw in ["Subgroup", "Secondary", "Post-Hoc"]):
                return "RCT_secundario", f"PubMed Publication Type contiene 'Randomized' + indicador de subanálisis ({type_str})"

        # Meta-análisis y revisiones sistemáticas
        if any(t in ("Meta-Analysis", "Systematic Review") for t in pub_types):
            return "meta-analisis", f"PubMed Publication Type = Meta-Analysis/Systematic Review ({type_str})"

        # Guías clínicas
        if any(t in ("Practice Guideline", "Guideline",
                     "Consensus Development Conference",
                     "Consensus Development Conference, NIH") for t in pub_types):
            return "guia_clinica", f"PubMed Publication Type = Guideline/Practice Guideline/Consensus ({type_str})"

        # ECA primario: requiere "Randomized Controlled Trial" Y que el
        # título/texto NO contenga señales de ser un análisis derivado
        # (subanálisis, subgrupo, post-hoc, secundario) que PubMed no
        # siempre marca explícitamente en sus Publication Types.
        if any(t == "Randomized Controlled Trial" for t in pub_types):
            derived_analysis_patterns = [
                r'\bsub.?group\b', r'\bsubanalysis\b', r'\bsub.?analysis\b',
                r'\bpost.?hoc\b', r'\bsecondary\s+analysis\b',
                r'\bsecondary\s+(endpoint|outcome)\b',
                r'\bpre.?specified\s+analysis\b', r'\bexploratory\s+analysis\b',
            ]
            check_text = " ".join(filter(None, [record.get("title", ""), record.get("ref_raw", "")]))
            for pat in derived_analysis_patterns:
                if re.search(pat, check_text, re.IGNORECASE):
                    return "RCT_secundario", f"PubMed = Randomized Controlled Trial, pero título/texto indica análisis derivado ('{pat}')"
            return "RCT_primario", f"PubMed Publication Type = Randomized Controlled Trial ({type_str})"

        # Observacional
        if any(t in ("Observational Study",) for t in pub_types):
            return "registro_observacional", f"PubMed Publication Type = Observational Study ({type_str})"

    # ── CAPA 2: Palabras clave en todos los campos de texto ──────────────
    text_to_search = " ".join(filter(None, [
        record.get("title", ""),
        record.get("ref_raw", ""),
        record.get("pub_type_raw", ""),
        record.get("journal", ""),
        " ".join(record.get("mesh_terms", [])),
    ]))

    for study_type in ["RCT_secundario", "meta-analisis", "guia_clinica",
                        "registro_observacional", "RCT_primario"]:
        for pat in CLASSIFICATION_RULES[study_type]:
            if re.search(pat, text_to_search, re.IGNORECASE):
                return study_type, f"Palabra clave en título/texto/journal/MeSH: patrón '{pat}'"

    # ── CAPA 3: Fallback por vocabulario de intervención ─────────────────
    # NOTA: esta capa es deliberadamente la de menor confianza. Solo debe
    # usarse cuando no hay pub_types ni ninguna keyword de capa 2, y aun
    # así requiere que el vocabulario de intervención aparezca junto a
    # una señal mínima de diseño comparativo (no basta con "effect of"
    # solo, que aparece igual en series de casos u observacionales).
    if record.get("pub_type_raw") == "journal-article":
        title = record.get("title", "")
        comparative_design_patterns = [
            r'\brandomi[sz]ed\b', r'\bplacebo\b', r'\btrial\b',
            r'\bdouble.blind\b', r'\bopen.label\b',
        ]
        has_comparative_signal = any(re.search(p, title, re.IGNORECASE) for p in comparative_design_patterns)
        if has_comparative_signal:
            for pat in [r'\beffect\s+of\b', r'\befficacy\b',
                        r'\bsafety\s+and\s+efficacy\b', r'\bversus\b',
                        r'\bcompar(ing|ison)\b', r'\bbenefit\s+of\b']:
                if re.search(pat, title, re.IGNORECASE):
                    return "RCT_primario", f"Sin pub_types/keywords; vocabulario de intervención ('{pat}') + señal de diseño comparativo en título"

    return "otro/no_clasificado", "No coincide con ningún patrón de las 3 capas (PubMed, keywords, vocabulario de intervención)"


# ─────────────────────────────────────────────
# MÓDULO 4: EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────

COLORS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "rct_primary": "E2EFDA",    # verde claro
    "rct_secondary": "FFF2CC",  # amarillo claro
    "meta": "DAE8FC",           # azul claro
    "registry": "F8CECC",       # rosa claro
    "guideline": "E1D5E7",      # lila claro
    "web_source": "FCE4D6",     # naranja claro
    "other": "F5F5F5",          # gris claro
    "subheader": "BDD7EE",
}

STUDY_TYPE_COLORS = {
    "RCT_primario": COLORS["rct_primary"],
    "RCT_secundario": COLORS["rct_secondary"],
    "meta-analisis": COLORS["meta"],
    "registro_observacional": COLORS["registry"],
    "guia_clinica": COLORS["guideline"],
    "fuente_web/institucional": COLORS["web_source"],
    "otro/no_clasificado": COLORS["other"],
}

def style_header(cell, bg_color=None, fg_color="FFFFFF", bold=True):
    bg = bg_color or COLORS["header_bg"]
    cell.font = Font(bold=bold, color=fg_color, name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_cell(cell, row_color=None, wrap=False):
    cell.font = Font(name="Arial", size=9)
    if row_color:
        cell.fill = PatternFill("solid", start_color=row_color)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)

def add_thin_border(ws, row, col_start, col_end):
    thin = Side(style="thin", color="CCCCCC")
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = Border(bottom=thin)


def export_to_excel(records: list[dict], output_path: str):
    wb = Workbook()

    # ── Hoja 1: Base de datos completa ──────────────────────────────
    ws_all = wb.active
    ws_all.title = "Todas las referencias"

    columns = [
        ("N°", 5), ("Autores", 30), ("Año", 6), ("Título", 50),
        ("Revista", 25), ("PMID", 12), ("DOI", 30),
        ("URL PubMed", 35), ("URL DOI", 35),
        ("Tipo (auto)", 20), ("Criterio", 45), ("Tipo (manual)", 20),
        ("Notas", 25), ("Abstract", 60), ("Referencia original", 50),
    ]

    # Encabezado
    ws_all.row_dimensions[1].height = 30
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws_all.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)
        ws_all.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws_all.freeze_panes = "A2"

    # Datos. Ordenamos por ref_number para que el Excel quede en orden
    # de lectura natural, pero usamos un contador de fila SECUENCIAL
    # (row_num = posición en la lista + 2) en vez de "ref_number + 1":
    # esto evita colisiones o huecos en el Excel si algún ref_number
    # no pudo extraerse correctamente (huecos, duplicados, etc.) — cada
    # referencia recibe siempre una fila propia, sin importar su número.
    sorted_records = sorted(records, key=lambda r: r.get("ref_number", 0))
    for row_offset, r in enumerate(sorted_records):
        row_num = row_offset + 2
        study_type = r.get("study_type_auto", "otro/no_clasificado")
        row_color = STUDY_TYPE_COLORS.get(study_type, COLORS["other"])

        values = [
            r.get("ref_number", ""),
            r.get("authors", ""),
            r.get("year", ""),
            r.get("title", ""),
            r.get("journal", ""),
            r.get("pmid", ""),
            r.get("doi", ""),
            r.get("pubmed_url", ""),
            r.get("doi_url", ""),
            r.get("study_type_auto", ""),
            r.get("classification_criterion", ""),
            r.get("study_type", ""),  # campo para corrección manual
            r.get("notes", ""),
            r.get("abstract", ""),
            r.get("ref_raw", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws_all.cell(row=row_num, column=col_idx, value=val)
            style_cell(cell, row_color=row_color, wrap=(col_idx in [4, 11, 14, 15]))
            # Hipervínculos
            if col_idx == 8 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            if col_idx == 9 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

        add_thin_border(ws_all, row_num, 1, len(columns))
        ws_all.row_dimensions[row_num].height = 40

    # ── Hoja 2: Solo ECAs primarios ─────────────────────────────────
    ws_rct = wb.create_sheet("ECAs primarios")
    rct_records = [r for r in records if r.get("study_type_auto") == "RCT_primario"]
    _write_rct_sheet(ws_rct, rct_records)

    # ── Hoja 3: Resumen por tipo ────────────────────────────────────
    ws_sum = wb.create_sheet("Resumen")
    _write_summary_sheet(ws_sum, records)

    # ── Hoja 4: Instrucciones ───────────────────────────────────────
    ws_help = wb.create_sheet("Instrucciones")
    _write_instructions_sheet(ws_help)

    wb.save(output_path)
    print(f"[OK] Excel guardado: {output_path}")


def _write_rct_sheet(ws, records):
    ws.title = "ECAs primarios"
    cols = [
        ("N°", 5), ("Autores", 30), ("Año", 6), ("Título", 50),
        ("Revista", 25), ("PMID", 12), ("DOI", 30),
        ("URL PubMed", 35), ("URL DOI", 35), ("Criterio", 45),
        ("Notas", 30), ("Abstract", 60),
    ]
    ws.row_dimensions[1].height = 30
    for col_idx, (col_name, col_width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell, bg_color="375623")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.freeze_panes = "A2"

    sorted_records = sorted(records, key=lambda r: r.get("ref_number", 0))
    for row_num, r in enumerate(sorted_records, 2):
        values = [
            r.get("ref_number"), r.get("authors"), r.get("year"),
            r.get("title"), r.get("journal"), r.get("pmid"), r.get("doi"),
            r.get("pubmed_url"), r.get("doi_url"), r.get("classification_criterion"),
            r.get("notes"), r.get("abstract"),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            style_cell(cell, row_color=COLORS["rct_primary"], wrap=(col_idx in [4, 10, 12]))
            if col_idx == 8 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
            if col_idx == 9 and val:
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
        ws.row_dimensions[row_num].height = 40


def _write_summary_sheet(ws, records):
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    headers = ["Tipo de estudio", "N referencias", "% del total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        style_header(cell)

    type_counts = {}
    for r in records:
        t = r.get("study_type_auto", "otro/no_clasificado")
        type_counts[t] = type_counts.get(t, 0) + 1

    total = len(records)
    for row_idx, (study_type, count) in enumerate(sorted(type_counts.items()), 2):
        pct = count / total * 100 if total else 0
        row_color = STUDY_TYPE_COLORS.get(study_type, COLORS["other"])
        ws.cell(row=row_idx, column=1, value=study_type).fill = PatternFill("solid", start_color=row_color)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(row=row_idx, column=3, value=f"{pct:.1f}%")

    # Total
    total_row = len(type_counts) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=2, value=total).font = Font(bold=True, name="Arial")
    ws.cell(row=total_row, column=3, value="100%").font = Font(bold=True, name="Arial")

    # Leyenda de colores
    ws.cell(row=total_row + 2, column=1, value="LEYENDA DE COLORES:").font = Font(bold=True, name="Arial", size=9)
    for idx, (st, color) in enumerate(STUDY_TYPE_COLORS.items(), total_row + 3):
        cell = ws.cell(row=idx, column=1, value=st)
        cell.fill = PatternFill("solid", start_color=color)
        cell.font = Font(name="Arial", size=9)


def _write_instructions_sheet(ws):
    ws.title = "Instrucciones"
    ws.column_dimensions["A"].width = 80

    lines = [
        ("GUÍA DE USO DEL PIPELINE DE REFERENCIAS", True, COLORS["header_bg"], "FFFFFF"),
        ("", False, None, None),
        ("MÓDULO 1 — Extracción del PDF", True, COLORS["subheader"], "000000"),
        ("El script localiza la ÚLTIMA aparición de 'References' en el documento y extrae", False, None, None),
        ("todo lo que sigue, usando la numeración secuencial de las referencias (no la", False, None, None),
        ("geometría de columnas) para reconstruir el texto correctamente en cualquier layout.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 2 — Enriquecimiento de metadatos", True, COLORS["subheader"], "000000"),
        ("Cada referencia se busca en CrossRef (texto libre) y PubMed (query por título+año).", False, None, None),
        ("Se extraen: PMID, DOI, título, autores, año, revista y ABSTRACT (resumen completo).", False, None, None),
        ("Se añaden URLs clicables a PubMed y DOI.", False, None, None),
        ("Las referencias claramente web/institucionales ('Available at: http://...') se", False, None, None),
        ("detectan y se excluyen de la búsqueda PubMed/CrossRef (no están indexadas allí).", False, None, None),
        ("NOTA: Sin API key de NCBI el límite es 3 req/s. El script respeta este límite.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 3 — Clasificación automática", True, COLORS["subheader"], "000000"),
        ("Clasificación en 3 capas de precisión decreciente:", False, None, None),
        ("  1) Publication Types oficiales de PubMed (más fiable)", False, None, None),
        ("  2) Palabras clave en título + texto crudo + journal + MeSH terms", False, None, None),
        ("  3) Vocabulario de intervención como último recurso", False, None, None),
        ("Categorías resultantes:", False, None, None),
        ("  • RCT_primario: ensayo clínico aleatorizado, publicación principal", False, None, None),
        ("  • RCT_secundario: subanálisis, post-hoc, subgrupos", False, None, None),
        ("  • meta-analisis: revisión sistemática, meta-análisis, pooled analysis", False, None, None),
        ("  • registro_observacional: registro, cohorte, observacional, retrospectivo", False, None, None),
        ("  • guia_clinica: guideline, consensus statement, task force, focused update", False, None, None),
        ("  • fuente_web/institucional: páginas web, recursos online sin DOI/PMID", False, None, None),
        ("  • otro/no_clasificado: no coincide con ningún patrón", False, None, None),
        ("La columna 'Tipo (manual)' permite correcciones manuales.", False, None, None),
        ("", False, None, None),
        ("MÓDULO 4 — Excel estructurado", True, COLORS["subheader"], "000000"),
        ("Hoja 'Todas las referencias': base de datos completa con código de colores por tipo,", False, None, None),
        ("incluye columna 'Abstract' con el resumen completo de PubMed cuando está disponible.", False, None, None),
        ("Hoja 'ECAs primarios': solo los ECAs primarios identificados, también con Abstract.", False, None, None),
        ("Hoja 'Resumen': tabla de frecuencias por tipo de estudio.", False, None, None),
        ("", False, None, None),
        ("USO EN LÍNEA DE COMANDOS", True, COLORS["subheader"], "000000"),
        ("  python guideline_pipeline.py guia.pdf output.xlsx", False, None, None),
        ("  python guideline_pipeline.py guia.pdf              (usa 'references_db.xlsx' por defecto)", False, None, None),
        ("", False, None, None),
        ("PARA GUÍAS CON API KEY DE NCBI (>3 req/s)", True, COLORS["subheader"], "000000"),
        ("  Añadir al entorno: export NCBI_API_KEY=tu_clave", False, None, None),
        ("  Obtener gratis en: https://www.ncbi.nlm.nih.gov/account/", False, None, None),
    ]

    for row_idx, (text, bold, bg, fg) in enumerate(lines, 1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(
            bold=bold, name="Arial", size=10,
            color=fg if fg else "000000"
        )
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_idx].height = 18


# ─────────────────────────────────────────────
# PIPELINE PRINCIPAL
# ─────────────────────────────────────────────

def run_pipeline(pdf_path: str, output_path: str = "references_db.xlsx"):
    print(f"\n{'='*60}")
    print(f"  PIPELINE DE REFERENCIAS - GUÍAS CLÍNICAS")
    print(f"{'='*60}")
    print(f"  PDF: {pdf_path}")
    print(f"  Output: {output_path}\n")

    # PASO 1: Extraer referencias
    print("[PASO 1] Extrayendo referencias del PDF...")
    raw_refs = extract_references_from_pdf(pdf_path)

    if not raw_refs:
        print("[ERROR] No se encontraron referencias. Verificar formato del PDF.")
        return

    # PASO 2 + 3: Enriquecer y clasificar
    print(f"\n[PASO 2+3] Enriqueciendo {len(raw_refs)} referencias con PubMed + CrossRef...")
    print("  (Esto puede tardar varios minutos para guías con muchas referencias)\n")

    records = []
    for idx, ref_text in enumerate(raw_refs, 1):
        print(f"  [{idx}/{len(raw_refs)}] {ref_text[:80]}...", end="\r")
        record = enrich_reference(ref_text, idx)
        record["study_type_auto"], record["classification_criterion"] = classify_reference(record)
        records.append(record)

    print(f"\n\n[INFO] Clasificación:")
    type_counts = {}
    for r in records:
        t = r["study_type_auto"]
        type_counts[t] = type_counts.get(t, 0) + 1
    for t, c in sorted(type_counts.items()):
        print(f"  {t}: {c}")

    # PASO 4: Exportar
    print(f"\n[PASO 4] Exportando a Excel: {output_path}")
    export_to_excel(records, output_path)

    print(f"\n{'='*60}")
    print(f"  COMPLETADO: {len(records)} referencias procesadas")
    print(f"  ECAs primarios identificados: {type_counts.get('RCT_primario', 0)}")
    print(f"{'='*60}\n")

    return records


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("USO: python guideline_pipeline.py <ruta_pdf> [output.xlsx]")
        sys.exit(1)

    pdf_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else "references_db.xlsx"
    run_pipeline(pdf_file, out_file)
